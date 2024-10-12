import telebot
from telebot.types import ReplyKeyboardMarkup, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardRemove
import requests
import logging
import mysql.connector
import DDL
import datetime
from datetime import datetime
from config import API_TOKEN, channel_cid, db_config
import openpyxl
import win32com.client
from managesalarytext import persian_text


bot = telebot.TeleBot(API_TOKEN)

hideboard = ReplyKeyboardRemove()

user_step = dict()

command = {"start": "getting started whit the telegram bot",
           "help": "show bot usage or frequently asked questions",
           "cancel": "canceling the operation"
           }


def listener(messages):
    for m in messages:
        if m.content_type == 'text':
            print(str(m.chat.first_name) + " [" + str(m.chat.id) + "]: " + m.text)
            logging.info(str(m.chat.first_name) + " [" + str(m.chat.id) + "]: " + m.text)


bot.set_update_listener(listener)


@bot.message_handler(commands=['start'])
def send_welcome(message):
    chat_name = message.chat.first_name
    cid = message.chat.id

    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor()

    SQL_QUERY = "SELECT cid, job_position FROM personnel_list"
    cursor.execute(SQL_QUERY)
    results = cursor.fetchall()
    cursor.close()
    conn.close()

    is_manager = False
    is_employee = False
    for row in results:
        db_cid, job_position = row
        if cid == db_cid and job_position == 'Manager':
            is_manager = True
            break
        elif cid == db_cid and job_position == 'Employee':
            is_employee = True
            break

    if len(message.text.split()) > 1:
            link = message.text.split()[-1]
            if link.startswith('new_employee'):
                enter_employee_by_link(message)
            elif link.startswith('new_manager'):
                enter_manager_by_link(message)
                return
    elif is_manager:
        bot.send_message(cid, f"{persian_text['welcome']}", parse_mode='markdown')
        reply_keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
        reply_keyboard.add(f"{persian_text['personnel_list']}", f"{persian_text['delete_employee']}",
                           "/cancel", f"{persian_text['invite_employee_link']}",f"{persian_text['invite_manager_link']}",
                           f"{persian_text['employee']}")
        bot.send_message(cid, f"{persian_text['manager_enter']}", reply_markup=reply_keyboard, parse_mode="markdown")
    elif is_employee:
        bot.send_message(cid, f"{persian_text['welcome']}", parse_mode='markdown')
        reply_keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
        reply_keyboard.add(f"{persian_text['enter_exit_hour']}", f"{persian_text['personal_info']}", f"{persian_text['change_picture']}",
                           f"{persian_text['manage_salary_excel']}", "/cancel",f"{persian_text['support']}")
        bot.send_message(cid, f"{persian_text['employee_enter']}", reply_markup=reply_keyboard, parse_mode="markdown")
    else:
        enter_employee_by_link(message)


@bot.message_handler(commands=["help"])
def help_command(message):
    cid = message.chat.id
    helpcommand = 13
    bot.copy_message(cid, channel_cid, helpcommand)


@bot.message_handler(commands=["cancel"])
def cancel_keyboard(message):
    cid = message.chat.id
    if user_step[cid] == "b" :
        reply_keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
        reply_keyboard.add(f"{persian_text['personnel_list']}", f"{persian_text['delete_employee']}",
                           "/cancel", f"{persian_text['invite_employee_link']}",
                           f"{persian_text['invite_manager_link']}",
                           f"{persian_text['employee']}")
        bot.send_message(cid, f"{persian_text['choose_keyboard']}",
                         reply_markup=reply_keyboard, parse_mode='markdown')
    else:
        bot.send_message(cid, f"{persian_text['cancel_process']}")
        user_step[cid] = None


@bot.message_handler(func=lambda m: m.text == f"{persian_text['invite_employee_link']}")
def invite_employee_link(message):
    cid = message.chat.id
    bot.send_message(cid, 'https://t.me/acounting_e_bot?start=new_employee')


@bot.message_handler(func=lambda m: m.text == "new_employee")
def enter_employee_by_link(message):
    cid = message.chat.id
    chat_name = message.chat.first_name
    bot.send_message(cid, f"Welcome {chat_name}!")
    bot.send_message(cid, f"{persian_text['send_your_data']} ")
    new_employee_format = 15
    bot.copy_message(cid, channel_cid, new_employee_format)
    user_step[cid] = "new_employee"


@bot.message_handler(func=lambda m: user_step.get(m.chat.id) == "new_employee")
def add_employee_by_link(message):
    cid = message.chat.id
    text = message.text
    employee_message = text.split('*')

    if len(employee_message) != 6:
        bot.send_message(cid, f"{persian_text['index_error']}")
        return

    try:
        name = employee_message[0].split(':')[-1].strip()[::-1]
        last_name = employee_message[1].split(':')[-1].strip()[::-1]
        job_position = "employee"
        rate = int(employee_message[2].split(':')[-1].strip())
        child_count = int(employee_message[3].split(':')[-1].strip())
        personnel_id = int(employee_message[4].split(':')[-1].strip())
        personnel_pass = int(employee_message[5].split(':')[-1].strip())
        picture_number = 16

        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        SQL_QUERY = """
            INSERT IGNORE INTO PERSONNEL_LIST 
            (cid, name, last_name, personnel_id, personnel_pass, job_position, child_count, rate, picture_number)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        cursor.execute(SQL_QUERY, (
            cid, name, last_name, personnel_id, personnel_pass, job_position, child_count, rate, picture_number))
        conn.commit()

        print(f"Employee {name + ' ' + last_name} added to the database.")
        bot.send_message(cid, f"{persian_text['adding_successfully']}")
        reply_keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
        reply_keyboard.add(f"{persian_text['enter_exit_hour']}", f"{persian_text['personal_info']}",
                           f"{persian_text['change_picture']}",
                           f"{persian_text['manage_salary_excel']}", "/cancel", f"{persian_text['support']}")
        notice = 31
        bot.copy_message(cid, channel_cid, notice)
        user_step[cid] = None

    except ValueError:
        bot.send_message(cid, f"{persian_text['wrong_info']}")

    except mysql.connector.Error as err:
        print(f"Error: {err}")
        bot.send_message(cid, f"{persian_text['database_error']}")

    finally:
        cursor.close()
        conn.close()


@bot.message_handler(func=lambda m: m.text == f"{persian_text['invite_manager_link']}")
def invite_manager_link(message):
    cid = message.chat.id
    bot.send_message(cid, 'https://t.me/acounting_e_bot?start=new_manager')


@bot.message_handler(func=lambda m: m.text == "new_manager")
def enter_manager_by_link(message):
    cid = message.chat.id
    chat_name = message.chat.first_name
    bot.send_message(cid, f"Welcome {chat_name}!")
    bot.send_message(cid, f"{persian_text['send_your_data']}")
    new_manager_format = 15
    bot.copy_message(cid, channel_cid, new_manager_format)
    user_step[cid] = "new_manager"


@bot.message_handler(func=lambda m: user_step.get(m.chat.id) == "new_manager")
def add_employee_by_link(message):
    cid = message.chat.id
    text = message.text
    employee_message = text.split('*')

    if len(employee_message) != 6:
        bot.send_message(cid, f"{persian_text['index_error']}")
        return

    try:
        name = employee_message[0].split(':')[-1].strip()[::-1]
        last_name = employee_message[1].split(':')[-1].strip()[::-1]
        job_position = "manager"
        rate = int(employee_message[2].split(':')[-1].strip())
        child_count = int(employee_message[3].split(':')[-1].strip())
        personnel_id = int(employee_message[4].split(':')[-1].strip())
        personnel_pass = int(employee_message[5].split(':')[-1].strip())
        picture_number = 16

        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        SQL_QUERY = """
            INSERT IGNORE INTO PERSONNEL_LIST 
            (cid, name, last_name, personnel_id, personnel_pass, job_position, child_count, rate, picture_number)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        cursor.execute(SQL_QUERY, (
            cid, name, last_name, personnel_id, personnel_pass, job_position, child_count, rate, picture_number))
        conn.commit()

        print(f"Employee {name + ' ' + last_name} added to the database.")
        bot.send_message(cid, f"{persian_text['adding_successfully']}")
        reply_keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
        reply_keyboard.add(f"{persian_text['enter_exit_hour']}", f"{persian_text['personal_info']}",
                           f"{persian_text['change_picture']}",
                           f"{persian_text['manage_salary_excel']}", "/cancel", f"{persian_text['support']}")
        notice = 31
        bot.copy_message(cid, channel_cid, notice)
        user_step[cid] = None

    except ValueError:
        bot.send_message(cid, f"{persian_text['wrong_info']}")

    except mysql.connector.Error as err:
        print(f"Error: {err}")
        bot.send_message(cid, f"{persian_text['database_error']}")

    finally:
        cursor.close()
        conn.close()


@bot.message_handler(func=lambda m: m.text == f"{persian_text['employee']}")
def keyboards_enter_employee(message):
    cid = message.chat.id
    reply_keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
    reply_keyboard.add(f"{persian_text['enter_exit_hour']}", f"{persian_text['personal_info']}",
                       f"{persian_text['change_picture']}",
                       f"{persian_text['manage_salary_excel']}", "/cancel", f"{persian_text['support']}")
    bot.send_message(cid, f"{persian_text['choose_keyboard']}", reply_markup=reply_keyboard, parse_mode='markdown')
    user_step[cid] =None


@bot.message_handler(func=lambda m: m.text == f"{persian_text['delete_employee']}")
def remove_employee_request(message):
    cid = message.chat.id
    bot.send_message(cid, f"{persian_text['enter_employee_id']}", parse_mode='markdown')
    notice = 32
    bot.copy_message(cid, channel_cid, notice, parse_mode='markdown')
    user_step[cid] = "a"


@bot.message_handler(func=lambda m: user_step.get(m.chat.id) == "a")
def remove_employee_response(message):
    cid = message.chat.id
    employee_cid = int(message.text)

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()

        SQL_QUERY_working_hours = """
                DELETE FROM working_hours 
                WHERE cid = %s
            """

        cursor.execute(SQL_QUERY_working_hours, (employee_cid,))
        print(f"employee with cid{employee_cid} delete from working hours")
        SQL_QUERY_timing ="""
            DELETE FROM timing 
                WHERE cid = %s
        """

        cursor.execute(SQL_QUERY_timing, (employee_cid,))
        print(f"employee with cid{employee_cid} delete from timing")

        SQL_QUERY_personnel_list = """
            DELETE FROM personnel_list 
                WHERE cid = %s
        """

        cursor.execute(SQL_QUERY_personnel_list, (employee_cid,))
        print(f"employee with cid{employee_cid} delete from personnel list")

        conn.commit()

        if cursor.rowcount > 0:
            bot.send_message(cid, f"{persian_text['delete_successfully']}", parse_mode='markdown')
        else:
            bot.send_message(cid, f"{persian_text['wrong_employee_data']}", parse_mode='markdown')

    except mysql.connector.Error as err:
        print(f"Error: {err}")
        bot.send_message(cid, f"{persian_text['database_error']}")

    finally:
        cursor.close()
        conn.close()
        user_step[cid] = "c"


@bot.message_handler(func=lambda m: m.text == f"{persian_text['personnel_list']}")
def list_of_employee(message):
    cid = message.chat.id
    inline_keyboard = InlineKeyboardMarkup()

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        SQL_QUERY = """
                SELECT name, last_name FROM personnel_list 

            """
        cursor.execute(SQL_QUERY)
        results = cursor.fetchall()
        for name, last_name in results:
            inline_keyboard.add(InlineKeyboardButton(name[::-1]+" "+last_name[::-1], callback_data=last_name))
        bot.send_message(cid,f"{persian_text['list_of_personnel']}", reply_markup=inline_keyboard, parse_mode='markdown')

    except mysql.connector.Error as err:
        print(f"Error:{err}")
        bot.send_message(cid, f"{persian_text['database_error']}")

    finally:
        cursor.close()
        conn.close()
    user_step[cid] = "b"


@bot.callback_query_handler(func=lambda call: user_step.get(call.message.chat.id) == "b")
def callback_query_function(call):
    cid = call.message.chat.id
    data = call.data
    call_id = call.id

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()

        SQL_QUERY = "SELECT * FROM PERSONNEL_LIST WHERE last_name = %s"
        cursor.execute(SQL_QUERY, (data,))
        info = cursor.fetchone()
        cursor.close()
        if info:
            response = (
                    f"{persian_text['full_name']} : {info[1][::-1]+' '+info[2][::-1]}\n"
                    f"{persian_text['cid']} : {info[0]}\n"  
                    f"{persian_text['password']} : {info[4]}\n"  
                    f"{persian_text['rate']} : {info[7]}\n"  
                    f"{persian_text['child_count']} : {info[6]}\n" 
                    f"{persian_text['personnel_id']} : {info[3]}\n"
                    f"{persian_text['job_position']} : {info[5]}\n"
                )

            bot.send_message(cid, response)
            reply_keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
            reply_keyboard.add(f"{persian_text['editing_personnel_id']}", f"{persian_text['editing_password']}"
                        ,f"{persian_text['editing_child_count']}", f"{persian_text['editing_rate']}","/cancel")
            bot.send_message(cid, f"{persian_text['notice_text']}",
                                reply_markup=reply_keyboard)


    except mysql.connector.Error as err:
        print(f"Error: {err}")
        bot.send_message(cid, f"{persian_text['database_error']}")

    finally:
        conn.close()


@bot.message_handler(func=lambda m: m.text == f"{persian_text['editing_personnel_id']}")
def edit_personnel_id_request(message):
    cid = message.chat.id
    sending_data = 33
    bot.copy_message(cid, channel_cid, sending_data, reply_markup=hideboard)
    user_step[cid] = "editing_personnel_id"


@bot.message_handler(func=lambda m: user_step.get(m.chat.id) == "editing_personnel_id")
def edit_personnelid_response(message):
    cid = message.chat.id
    text = message.text.split('_')
    old_personnel_id = int(text[0].strip())
    new_personnel_id= int(text[1].strip())

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        SQL_QUERY = """
           update personnel_list set personnel_id = %s where personnel_id = %s

            """
        cursor.execute(SQL_QUERY,(new_personnel_id,old_personnel_id))
        conn.commit()
        results = cursor.fetchall()
        reply_keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
        reply_keyboard.add(f"{persian_text['personnel_list']}", f"{persian_text['delete_employee']}",
                           "/cancel", f"{persian_text['invite_employee_link']}",
                           f"{persian_text['invite_manager_link']}",
                           f"{persian_text['employee']}")
        bot.send_message(cid, f"{persian_text['editing_personnel_id_successfully']}",reply_markup=reply_keyboard)
        user_step[cid] = None
    except mysql.connector.Error as err:
        print(f"Error: {err}")
        bot.send_message(cid, f"{persian_text['database_error']}")
    finally:
        cursor.close()
        conn.close()


@bot.message_handler(func=lambda m: m.text == f"{persian_text['editing_password']}")
def edit_pass_request(message):
    cid = message.chat.id
    sending_data = 34
    bot.copy_message(cid, channel_cid, sending_data, reply_markup=hideboard)
    user_step[cid] = "editing_password"


@bot.message_handler(func=lambda m: user_step.get(m.chat.id) == "editing_password")
def edit_personnel_pass_response(message):
    cid = message.chat.id
    text = message.text.split('_')
    personnel_id = int(text[0].strip())
    new_personnel_pass= int(text[1].strip())

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        SQL_QUERY = """
           update personnel_list set personnel_pass = %s where personnel_id = %s

            """
        cursor.execute(SQL_QUERY,(new_personnel_pass,personnel_id))
        conn.commit()
        results = cursor.fetchall()
        reply_keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
        reply_keyboard.add(f"{persian_text['personnel_list']}", f"{persian_text['delete_employee']}",
                           "/cancel", f"{persian_text['invite_employee_link']}",
                           f"{persian_text['invite_manager_link']}",
                           f"{persian_text['employee']}")
        bot.send_message(cid, f"{persian_text['editing_password_successfully']}", reply_markup=reply_keyboard)
        user_step[cid] = None
    except mysql.connector.Error as err:
        print(f"Error: {err}")
        bot.send_message(cid, f"{persian_text['database_error']}")
        user_step[cid] = None
    finally:
        cursor.close()
        conn.close()


@bot.message_handler(func=lambda m: m.text == f"{persian_text['editing_child_count']}")
def edit_child_count_request(message):
    cid = message.chat.id
    sending_data = 35
    bot.copy_message(cid, channel_cid, sending_data, reply_markup=hideboard)
    user_step[cid] = "editing_child_count"


@bot.message_handler(func=lambda m: user_step.get(m.chat.id) == "editing_child_count")
def edit_child_count_response(message):
    cid = message.chat.id
    text = message.text.split('_')
    personnel_id = int(text[0].strip())
    new_child_count= int(text[1].strip())

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        SQL_QUERY = """
           update personnel_list set child_count = %s where personnel_id = %s

            """
        cursor.execute(SQL_QUERY,(new_child_count,personnel_id))
        conn.commit()
        results = cursor.fetchall()
        reply_keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
        reply_keyboard.add(f"{persian_text['personnel_list']}", f"{persian_text['delete_employee']}",
                           "/cancel", f"{persian_text['invite_employee_link']}",
                           f"{persian_text['invite_manager_link']}",
                           f"{persian_text['employee']}")
        bot.send_message(cid, f"{persian_text['editing_child_count_successfully']}", reply_markup=reply_keyboard)
        user_step[cid] = None
    except mysql.connector.Error as err:
        print(f"Error: {err}")
        bot.send_message(cid, f"{persian_text['database_error']}")
    finally:
        cursor.close()
        conn.close()


@bot.message_handler(func=lambda m: m.text == f"{persian_text['editing_rate']}")
def edit_employee_rate_request(message):
    cid = message.chat.id
    sending_data = 36
    bot.copy_message(cid, channel_cid, sending_data, reply_markup=hideboard)
    user_step[cid] = "editing_rate"


@bot.message_handler(func=lambda m: user_step.get(m.chat.id) == "editing_rate")
def edit_rate_response(message):
    cid = message.chat.id
    text = message.text.split('_')
    personnel_id = int(text[0].strip())
    new_rate= int(text[1].strip())

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        SQL_QUERY = """
           update personnel_list set rate = %s where personnel_id = %s

            """
        cursor.execute(SQL_QUERY,(new_rate, personnel_id))
        conn.commit()
        results = cursor.fetchall()
        reply_keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
        reply_keyboard.add(f"{persian_text['personnel_list']}", f"{persian_text['delete_employee']}",
                           "/cancel", f"{persian_text['invite_employee_link']}",
                           f"{persian_text['invite_manager_link']}",
                           f"{persian_text['employee']}")
        bot.send_message(cid, f"{persian_text['editing_rate_successfully']}",reply_markup=reply_keyboard)
        user_step[cid] = None

    except mysql.connector.Error as err:
        print(f"Error: {err}")
        bot.send_message(cid, f"{persian_text['database_error']}")

    finally:
        cursor.close()
        conn.close()


@bot.message_handler(func=lambda m: m.text ==f"{persian_text['enter_exit_hour']}")
def enter_exit_time_keyboard(message):
    cid = message.chat.id
    sending_data = 37
    bot.copy_message(cid, channel_cid, sending_data)
    enter_exit_format = 38
    bot.copy_message(cid, channel_cid, enter_exit_format)
    user_step[cid] = 10


@bot.message_handler(func=lambda m: user_step.get(m.chat.id) == 10)
def enter_exit_time_request(message):
    cid = message.chat.id
    enter_exit_time = message.text.split("_")
    enter_time = enter_exit_time[0]
    exit_time = enter_exit_time[1]
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        check_query = """
                    SELECT * FROM TIMING
                    WHERE cid = %s AND date = CURDATE()
                """
        cursor.execute(check_query, (cid,))
        existing_entry = cursor.fetchone()

        if existing_entry:
            bot.send_message(cid, f"{persian_text['timing_error']}")
            user_step[cid] = None
        else:
            cursor.execute("SELECT cid FROM personnel_list WHERE cid = %s", (cid,))
            database_cid = cursor.fetchone()
            if database_cid:
                database_cid = database_cid[0]
                print(f"{database_cid} entered working hours time.")
            else:
                raise ValueError("cid not found")

            SQL_QUERY = """
               INSERT IGNORE INTO TIMING
               (`cid`, `date`, `start_time`, `end_time`)
               VALUES (%s,CURDATE(), %s, %s );
           """
            cursor.execute(SQL_QUERY, (database_cid, enter_time, exit_time))
            conn.commit()
            today_date = datetime.now()
            formatted_date = today_date.strftime('%d-%B')
            message = f"{persian_text['enter_exit_successfully']}"
            bot.send_message(cid, f" {formatted_date}\n{message}")
            user_step[cid] = None

    except mysql.connector.Error as err:
        print(f"Error: {err}")
        bot.send_message(cid, f"{persian_text['database_error']}")


    finally:
        cursor.close()
        conn.close()
        insert_working_hours(cid)


@bot.message_handler(func=lambda m: m.text == f"{persian_text['change_picture']}")
def changing_personnel_picture(message):
    cid = message.chat.id
    bot.send_message(cid, f"{persian_text['sending_picture']}")
    user_step[cid] = 20


@bot.message_handler(func=lambda m: True, content_types=['photo'])
def personnel_picture(message):
    cid = message.chat.id
    if user_step[cid]==20:
        file_id = message.photo[-1].file_id
        res = bot.send_photo(channel_cid, file_id)
        channel_mid = res.message_id

        try:
            conn = mysql.connector.connect(**db_config)
            cursor = conn.cursor()

            SQL_QUERY = """
                        update personnel_list set picture_number = %s where cid = %s
                    """
            cursor.execute(SQL_QUERY, (channel_mid,cid))
            conn.commit()
            info = cursor.fetchone()
            cursor.close()
            bot.send_message(cid, f"{persian_text['changing_picture_successfully']}")
            user_step[cid] = None

        except mysql.connector.Error as err:
            print(f"Error: {err}")
            bot.send_message(cid, f"{persian_text['database_error']}")

        finally:
            conn.close()
    else:
        bot.send_message(cid,f"{persian_text['invalid_message']}")


@bot.message_handler(func=lambda m: m.text ==f"{persian_text['personal_info']}")
def personal_information_request(message):
    cid = message.chat.id
    bot.send_message(cid, f"{persian_text['enter_password']}", parse_mode='markdown')
    user_step[cid] = 30


@bot.message_handler(func=lambda m: user_step.get(m.chat.id) == 30)
def personal_information_response(message):
    cid = message.chat.id
    employee_pass = int(message.text)

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()

        SQL_QUERY = """
                    SELECT * FROM PERSONNEL_LIST 
                    WHERE cid = %s AND personnel_pass = %s
                """
        cursor.execute(SQL_QUERY, (cid, employee_pass))
        info = cursor.fetchone()
        cursor.close()


        if info:
            response = (
                f"{persian_text['full_name']} : {info[1][::-1] + ' ' + info[2][::-1]}\n"
                f"{persian_text['password']} : {info[4]}\n"
                f"{persian_text['rate']} : {info[7]}\n"
                f"{persian_text['child_count']} : {info[6]}\n"
                f"{persian_text['personnel_id']} : {info[3]}\n"
                f"{persian_text['job_position']} : {info[5]}\n"
            )

            bot.copy_message(cid, channel_cid, info[8], caption=response)
            reply_keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
            reply_keyboard.add(f"{persian_text['editing_name']}", f"{persian_text['editing_lastname']}",
                               f"{persian_text['cancel_editing']}")
            bot.send_message(cid, f"{persian_text['notice_text']}",
                             reply_markup=reply_keyboard)
            user_step[cid] = None
        else:
            bot.send_message(cid, f"{persian_text['wrong_password']}")
            user_step[cid] = None

    except mysql.connector.Error as err:
        print(f"Error: {err}")
        bot.send_message(cid, f"{persian_text['database_error']}")

    finally:
        conn.close()

@bot.message_handler(func=lambda m: m.text == f"{persian_text['cancel_editing']}")
def edit_employee_name_request(message):
    cid = message.chat.id
    reply_keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
    reply_keyboard.add(f"{persian_text['enter_exit_hour']}", f"{persian_text['personal_info']}",
                       f"{persian_text['change_picture']}",
                       f"{persian_text['manage_salary_excel']}", "/cancel", f"{persian_text['support']}")
    bot.send_message(cid, f"{persian_text['choose_keyboard']}",
                     reply_markup=reply_keyboard, parse_mode='markdown')


@bot.message_handler(func=lambda m: m.text == f"{persian_text['editing_name']}")
def edit_employee_name_request(message):
    cid = message.chat.id
    bot.send_message(cid, f"{persian_text['enter_new_name']}",reply_markup=hideboard)
    user_step[cid] = 'editting_name'


@bot.message_handler(func=lambda m: user_step.get(m.chat.id) == "editting_name")
def edit_employee_name_response(message):
    cid = message.chat.id
    new_name = str(message.text[::-1])

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        SQL_QUERY = """
           update personnel_list set name = %s where cid = %s

            """
        cursor.execute(SQL_QUERY,(new_name,cid))
        conn.commit()
        results = cursor.fetchall()
        reply_keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
        reply_keyboard.add(f"{persian_text['enter_exit_hour']}", f"{persian_text['personal_info']}",
                           f"{persian_text['change_picture']}",
                           f"{persian_text['manage_salary_excel']}", "/cancel", f"{persian_text['support']}")
        bot.send_message(cid, f"{persian_text['editing_name_successfully']}",reply_markup=reply_keyboard)
        user_step[cid] = None

    except mysql.connector.Error as err:
        print(f"Error: {err}")
        bot.send_message(cid, f"{persian_text['database_error']}")

    finally:
        cursor.close()
        conn.close()


@bot.message_handler(func=lambda m: m.text == f"{persian_text['editing_lastname']}")
def edit_employee_lastname_request(message):
    cid = message.chat.id
    bot.send_message(cid, f"{persian_text['enter_new_lastname']}", reply_markup=hideboard)
    user_step[cid] = 'editing_lastname'


@bot.message_handler(func=lambda m: user_step.get(m.chat.id) == "editing_lastname")
def edit_employee_lastname_response(message):
    cid = message.chat.id
    new_lastname = str(message.text[::-1])
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        SQL_QUERY = f"""
          update personnel_list set last_name = %s where cid = %s

            """
        cursor.execute(SQL_QUERY,(new_lastname,cid))
        conn.commit()
        results = cursor.fetchall()
        reply_keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
        reply_keyboard.add(f"{persian_text['enter_exit_hour']}", f"{persian_text['personal_info']}",
                           f"{persian_text['change_picture']}",
                           f"{persian_text['manage_salary_excel']}", "/cancel", f"{persian_text['support']}")
        bot.send_message(cid, f"{persian_text['editing_lastname_successfully']}", reply_markup=reply_keyboard)
        user_step[cid] = None

    except mysql.connector.Error as err:
        print(f"Error: {err}")
        bot.send_message(cid, f"{persian_text['database_error']}")

    finally:
        cursor.close()
        conn.close()


@bot.message_handler(func=lambda m: m.text == f"{persian_text['support']}")
def supporting_request(message):
    cid = message.chat.id
    telegram_cid_support = "[support](tg://user?id=667547362"
    bot.send_message(cid,f"{persian_text['press_to_support']} : {telegram_cid_support}", parse_mode='markdown')

@bot.message_handler(func=lambda m: m.text == f"{persian_text['manage_salary_excel']}")
def excel_managesalary(message):
    cid = message.chat.id
    calc = CalculationPayment()
    calc.getting_info(message)
    calc.insert_excel_role_info(message)
    calc.insert_excel_income_calculation(message)
    calc.insert_excel_deficit_calculation(message)
    calc.insert_excel_payment_calculation(message)
    today_date = datetime.now()
    formatted_date = today_date.strftime('%d-%B')

    with open("managesalaryexcel.xlsx","rb") as f:
        bot.send_document(cid, f)
    text = f"{persian_text['calculation_salary_successfully']}"
    bot.send_message(cid,f"{formatted_date}\n {text}")


class CalculationPayment:
    def getting_info(self, message):
        self.cid = message.chat.id
        try:
            conn = mysql.connector.connect(**db_config)
            cursor = conn.cursor()
            current_date = datetime.now()
            current_month = current_date.month

            SQL_QUERY_working_hours = """
                SELECT * FROM working_hours 
                WHERE cid = %s AND month = %s
            """
            cursor.execute(SQL_QUERY_working_hours, (self.cid, current_month))
            role = cursor.fetchone()
            self.working_hour = role[2]

            SQL_QUERY_personnel_list = """
                SELECT * FROM personnel_list 
                WHERE cid = %s 
            """
            cursor.execute(SQL_QUERY_personnel_list, (self.cid,))
            info = cursor.fetchone()
            self.name = info[1]
            self.last_name = info[2]
            self.personnel_id = info[3]
            self.child_count = info[6]
            self.rate = info[7]

        except mysql.connector.Error as err:
            print(f"Error: {err}")

        finally:
            cursor.close()
            conn.close()

    def calculation_base_payment(self, message):
        base_payment = self.working_hour * self.rate
        return base_payment

    def calculation_child_subsidy(self, message):
        child_subsidy = self.rate * self.child_count * 3
        return child_subsidy

    def calculation_payment_benefits(self, message):
        payment_benefits  = 900000 + 500000
        return payment_benefits

    def calculation_overtime_benefits(self, message):
        if self.working_hour > 176 :
            overtime_payment = (self.working_hour - 176) * (self.rate * 140) / 100
            return overtime_payment
        return 0

    def insert_excel_role_info(self,message):
        wb = openpyxl.load_workbook("managesalaryexcel.xlsx")
        sheet = wb.active
        today = datetime.now()

        sheet["B2"] = str(self.name[::-1])+" "+str(self.last_name[::-1])
        sheet["D2"] = self.personnel_id
        sheet["F2"] = today.strftime("%Y/%B/%d")
        sheet["B5"] = self.working_hour
        if self.working_hour>176:
            sheet["B6"] = int(self.working_hour)-176
        else:
            sheet["B6"] = 0
        sheet["B7"] = self.child_count
        sheet["B8"] = self.rate

        wb.save("managesalaryexcel.xlsx")
        wb.close()
        return

    def insert_excel_income_calculation(self,message):
        wb = openpyxl.load_workbook("managesalaryexcel.xlsx")
        sheet = wb.active

        sheet["D5"] = self.working_hour*self.rate
        sheet["D6"] = f"=PRODUCT(B6,{self.rate})*140%"
        sheet["D7"] = self.child_count*self.rate*3
        sheet["D8"] = 500000
        sheet["D9"] = 900000
        sheet["D10"] = "=SUM(D5,D6,D7,D8,D9)"

        wb.save("managesalaryexcel.xlsx")
        wb.close()
        return


    def insert_excel_deficit_calculation(self,message):
        wb = openpyxl.load_workbook("managesalaryexcel.xlsx")
        sheet = wb.active

        sheet["F5"] = "=SUM(D5,D8,D9)*7%"
        sheet["F6"] = "=D10*10%"
        sheet["F7"] = "=SUM(F5,F6)"

        wb.save("managesalaryexcel.xlsx)
        wb.close()
        return

    def insert_excel_payment_calculation(self, message):
        wb = openpyxl.load_workbook("managesalaryexcel.xlsx")
        sheet = wb.active

        sheet["F10"] = "=D10-F7"

        wb.save("managesalaryexcel.xlsx")
        wb.close()
        return


def insert_working_hours(cid):
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor(dictionary=True)

    SQL_QUERY_GET = """
            SELECT cid, MONTH(date) AS month, SUM(TIMESTAMPDIFF(MINUTE, start_time, end_time)) / 60 AS working_hours
            FROM TIMING
            WHERE cid = %s
            GROUP BY cid, month ;"""

    cursor.execute(SQL_QUERY_GET, (cid,))
    working_hours_data = cursor.fetchall()

    SQL_QUERY_INSERT = """
            INSERT INTO working_hours (cid, month, working_hours)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE working_hours = VALUES(working_hours) ;"""

    for row in working_hours_data:
        cursor.execute(SQL_QUERY_INSERT, (row['cid'], row['month'], row['working_hours']))

    conn.commit()
    cursor.close()
    conn.close()


bot.infinity_polling()




